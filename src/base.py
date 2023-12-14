import database.database as db
import numpy as npy
import matplotlib.pyplot as plt
import openpyxl as xl
import pandas as pd
import math


def show_img(id, table):
    """
    根据指定 ID，寻找数据库中的制定图片并展示出来。
    :param id: 图片 ID
    :param table: 所要寻找的数据表
    :return: 无返回值
    """
    # 获取数据
    img = table.select_where_id(id)
    img_data = npy.frombuffer(img[0][1], dtype=npy.uint8)
    img_data.resize((28, 28))

    # 开始绘制
    plt.imshow(255-img_data, cmap='gray', vmin=0, vmax=255)
    plt.colorbar()
    plt.show()


class Sigmoid:
    @staticmethod
    def f(x):
        return 1 / (1 + npy.exp(-x))


    @staticmethod
    def f_p(x):
        return Sigmoid.f(x) * (1 - Sigmoid.f(x))


class NeuralLayer:
    """
    神经层
    """
    def __init__(self, last_total=1, current_total=1, activation_fun=Sigmoid):
        """
        :param last_total    前一层的神经元的总数量
        :param current_total 这一层的神经元的总数量
        """
        self.last_total = last_total
        self.current_total = current_total

        # 激活函数
        self.activation_fun = activation_fun


        # A：激活值表
        # Z：中间值表
        # A_last：上一层的激活值，也就是输入值
        # B：偏置表
        # W：权重表
        self._A = npy.zeros(current_total, dtype=npy.float32)
        self._Z = npy.zeros(current_total, dtype=npy.float32)
        self._A_last = npy.zeros(last_total, dtype=npy.float32)
        self._B = npy.zeros(current_total, dtype=npy.float32)
        # self._W = npy.zeros((current_total, last_total), dtype=npy.float32)

        self._W = npy.array(npy.random.random((current_total, last_total)))


    def execute(self, A_last: npy.ndarray):
        """
        根据给定的前一层神经元的激活值改变该层神经元的激活值
        :param A_last: 前一层神经元的激活值
        :return: 无返回值
        """
        self._A_last = A_last
        self._Z = self._W @ self._A_last + self._B
        self._A = self.activation_fun.f(self._Z)


    def backpropagation(self, delta: npy.ndarray):
        """
        反向传播算法更新 W、B
        :param delta: 这一层的 delta 向量
        :return: 返回后面一层的 delta 向量
        """
        # delta_base = self._A_last.reshape((1, -1)).repeat(self.current_total, axis=0)
        # sigmoid_p_Z = sigmoid_partial(self._Z)
        #
        # delta_B = sigmoid_p_Z * (2 * delta)
        # delta_W = (delta_base.T * delta_B).T
        #
        # delta_A = npy.array([sum(self._W[:,i] * delta_B) for i in range(self._W.shape[1])])
        #
        # self._B -= delta_B
        # self._W -= delta_W
        #
        # return delta_A

        sigmoid_p_Z = self.activation_fun.f_p(self._Z)

        delta_B = sigmoid_p_Z * (2 * delta)
        delta_W = self._A_last * delta_B.reshape((-1, 1))
        delta_A = (delta_W.T @ delta_B.reshape((-1, 1))).sum(axis=1)

        self._B -= delta_B
        self._W -= delta_W
        return delta_A


    def set_A(self, A: npy.ndarray):
        if len(A) != len(self._A):
            print("NeuralLayer::set_A(): 尺寸不相符！")
        self._A = npy.array(A)


    def load(self, sheet):
        """
        读取数据表中的数据
        :param sheet: 工作表对象。注意不是传入excel，而是sheet
        """
        data = [[cell.value for cell in row] for row in sheet.rows]
        self._W = npy.array([row for row in data[:-1]], dtype=npy.float32)
        self._B = npy.array(data[-1], dtype=npy.float32)
        self._B = self._B[npy.flatnonzero(~npy.isnan(self._B))]  # 这一步是去除 None

        self.current_total = self._W.shape[0]
        self.last_total = self._W.shape[1]


    def save(self, sheet):
        """
        将该层神经网络存储到对于的 excel 表格中。
        在将数据存入到 excel 中，为了方便计算，写入的顺序是：W、B ！！！
        :param sheet: 工作表对象。注意不是传入excel，而是sheet
        """
        for item in self._W:
            sheet.append(list(item))
        sheet.append(list(self._B))


class NeuralNetwork:
    def __init__(self):
        self.size = [1, 28*28, 16, 16, 10]
        self.network = [NeuralLayer(self.size[i], self.size[i+1]) for i in range(len(self.size)-1)]


    def execute(self, input):
        """
        根据输入的参数，运行神经网络。
        注意：输入值应该是一个竖向量
        :param input: 输入参数，这是一个竖向量
        """
        self.network[0].set_A(input)
        for i in range(1, len(self.network)):
            self.network[i].execute(self.network[i-1]._A)


    def backpropagation(self, graph, target):
        """
        训练神经网络
        :return:
        """
        self.execute(graph)
        result = [0] * 10
        result[target] = 1

        delta = self.network[-1]._A - result
        for i in range(len(self.network)-1, 0, -1):
            delta = self.network[i].backpropagation(delta)


    def debug(self):
        """
        打印神经网络的结果
        """
        for i in range(self.size[-1]):
            print(f'{i} -> {self.network[-1]._A[i]:.5f}')


    def save(self, file):
        """
        将该层神经网络存储到对于的 excel 中。
        :param file: 文件路径。
        """
        wb = xl.Workbook()

        number = 1
        for layer in self.network[1:]:
            ws = wb.create_sheet(f'layer{number}')
            layer.save(ws)
            number += 1

        # wb.remove('Sheet')
        wb.save(file)


    def load(self, file):
        """
        将神经网络从 excel 中读取出来
        :param file: 文件路径
        """
        wb = xl.load_workbook(file)

        for i in range(1, len(wb.sheetnames)):
            self.network[i].load(wb[f'layer{i}'])



